import psycopg2
import psycopg2.extras
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side

# Database configurations
DB_CONFIG = {
    'host': 'delinquently-sturdy-kelpie.data-1.use1.tembo.io',
    'user': 'postgres',
    'password': 'gO4BrtT9Z0tTyqBA',
    'dbname': 'postgres'
}

def connect_to_db():
    return psycopg2.connect(**DB_CONFIG)

def create_database_and_table():
    conn = connect_to_db()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS rotas (
            id SERIAL PRIMARY KEY,
            rota VARCHAR(10) NOT NULL,
            doca INT NOT NULL,
            observacao VARCHAR(200)
        )
    """)
    conn.commit()
    cursor.close()
    conn.close()

def insert_data(rota, doca, observacao):
    conn = connect_to_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO rotas (rota, doca, observacao) VALUES (%s, %s, %s)",
                   (rota, doca, observacao))
    conn.commit()
    cursor.close()
    conn.close()

def query_data(id=None, rota=None, doca=None):
    conn = connect_to_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    query = "SELECT * FROM rotas WHERE 1=1"
    params = []
    if id:
        query += " AND id = %s"
        params.append(id)
    if rota:
        query += " AND rota = %s"
        params.append(rota)
    if doca:
        query += " AND doca = %s"
        params.append(doca)
    query += " ORDER BY rota ASC"
    cursor.execute(query, params)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results

def delete_data(id=None, rota=None, doca=None):
    conn = connect_to_db()
    cursor = conn.cursor()
    query = "DELETE FROM rotas WHERE 1=1"
    params = []
    if id:
        query += " AND id = %s"
        params.append(id)
    if rota:
        query += " AND rota = %s"
        params.append(rota)
    if doca:
        query += " AND doca = %s"
        params.append(doca)
    cursor.execute(query, params)
    conn.commit()
    cursor.close()
    conn.close()

def export_to_excel(data):
    df = pd.DataFrame(data, columns=["id", "rota", "doca", "observacao"])
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

    excel_buffer.seek(0)  # Reset buffer position
    return excel_buffer

def main():
    create_database_and_table()
    st.title("  Onde Está a Rota")

    # Form to add new route
    with st.expander("Adicionar Nova Rota"):
        with st.form("add_form", clear_on_submit=True):
            rota = st.text_input("Rota")
            doca = st.number_input("Doca", min_value=1)
            observacao = st.text_area("Observação")
            submit_button = st.form_submit_button("Adicionar")
            if submit_button:
                insert_data(rota, doca, observacao)
                st.success("Rota adicionada com sucesso!")

    # Form to query routes
    with st.expander("Consultar Rota"):
        with st.form("search_form", clear_on_submit=True):
            id_search = st.number_input("ID", min_value=0, step=1)
            rota_search = st.text_input("Rota (opcional)")
            doca_search = st.number_input("Doca (opcional)", min_value=0, step=1)
            search_button = st.form_submit_button("Consultar")
            results = []
            if search_button:
                results = query_data(
                    id_search if id_search else None,
                    rota_search or None,
                    doca_search if doca_search else None
                )
                if results:
                    df = pd.DataFrame(results, columns=["id", "rota", "doca", "observacao"])
                    st.write(df)
                else:
                    st.warning("Nenhuma rota encontrada.")
        # Display download button outside the form
        if results:
            excel_data = export_to_excel(results)
            st.download_button(
                label="Baixar Excel",
                data=excel_data,
                file_name="consultas_rotas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # Form to delete route
    with st.expander("Excluir Rota"):
        with st.form("delete_form", clear_on_submit=True):
            id_delete = st.number_input("ID para excluir", min_value=0, step=1)
            rota_delete = st.text_input("Rota para excluir (opcional)")
            doca_delete = st.number_input("Doca para excluir (opcional)", min_value=0, step=1)
            delete_button = st.form_submit_button("Excluir")
            if delete_button:
                delete_data(
                    id_delete if id_delete else None,
                    rota_delete or None,
                    doca_delete if doca_delete else None
                )
                st.success("Rota excluída com sucesso!")

if __name__ == "__main__":
    main()
